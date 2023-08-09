import json
import re

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from googletrans import Translator
from re import search
from datetime import datetime

import importlib.util

from config.colors import *
from config.config import *
from config.materials import *

from random import random
import shutil
import requests
import os
import logging
import time
import sys
import traceback
import argparse


class Parser:
    def __init__(self):
        self.result = []
        parser = argparse.ArgumentParser(description='Process some integers.')
        parser.add_argument('--headless', action='store_true', help='headless')
        args = parser.parse_args()
        if args.headless:
            self.driver = self.get_driver(True)
        else:
            self.driver = self.get_driver(False)

    def get_driver(self, headless):
        try:
            options = webdriver.ChromeOptions()
            if headless:
                options.add_argument('--headless')
                options.add_argument('--disable-gpu')

            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)

            options.add_argument(
                "user-agent=Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:84.0) Gecko/20100101 Firefox/84.0")

            # options.add_argument('--disable-dev-shm-usage')
            # options.add_argument('--no-sandbox')

            driver = webdriver.Chrome(
                service=Service('chromedriver.exe'),
                options=options
            )
            driver.set_window_size(1920, 1080)
            driver.implicitly_wait(30)

            self.wait = WebDriverWait(driver, 30)

            return driver
        except Exception as e:
            print('Неудачная настройка браузера!')
            print(traceback.format_exc())
            print(input('Нажмите ENTER, чтобы закрыть эту программу'))
            sys.exit()

    def get_all_products(self):
        products = []
        # scroll page
        while self.check_exists_by_xpath('//button[text()="Next Page"]'):
            self.driver.execute_script(f"window.scrollTo(0, 10000);")
            products.extend([i.get_attribute('href') for i in self.driver.find_elements(By.XPATH, '//a[@data-ticket]')])
            btn = self.driver.find_element(By.XPATH, '//nav[@aria-label="Pagination"]/ul/li[last()]/button')
            btn.click()
            time.sleep(TIMEOUT)
        products.extend([i.get_attribute('href') for i in self.driver.find_elements(By.XPATH, '//a[@data-ticket]')])

        return list(set(products))

    def parse(self):
        c = 0
        self.driver.get(self.CATEGORIE_URL)
        time.sleep(TIMEOUT)

        products = self.get_all_products()

        for product_url in products[:PARSE_LIMIT]:
            print(f'{products.index(product_url) + 1} of {len(products[:PARSE_LIMIT])}')
            try:
                self.driver.get(product_url)
            except:
                continue

            name = self.translate(self.driver.find_element(By.XPATH, '//h1').text)

            description = self.driver.find_element(By.XPATH, '//div[@id="description"]').text

            material = []
            for word in description.split('% '):
                for i in self.MATERIALS.keys():
                    if word.lower().startswith(i):
                        material.append(self.MATERIALS[i])
                        break
            material = ';'.join(material)

            self.driver.find_element(By.CLASS_NAME, 'a-link.details').click()
            time.sleep(TIMEOUT)
            self.driver.find_element(By.CLASS_NAME, 'a-link.open-lightbox.size-guide').click()
            time.sleep(TIMEOUT)
            creator = self.driver.find_element(By.CLASS_NAME, 'country-name-value').text
            self.driver.find_element(By.CLASS_NAME, 'a-button-nostyle.m-button-icon').click()
            if creator == '':
                print(self.driver.current_url)

            eur_price = self.driver.find_element(By.CLASS_NAME, 'price').text.replace('€', '').replace(',', '.').strip()
            price = self.get_price(eur_price)

            btn = self.driver.find_element(By.ID, 'pdp-dropdown-label')
            btn.click()
            color = btn.text.strip()
            self.driver.execute_script("window.scrollTo(0, 10000);")


            article_num = re.search('[0-9]{5,}', self.driver.current_url)[0]

            photos = self.driver.find_elements(By.XPATH, '//img[contains(@id, "gallery-product")]')
            main_photo_url = 'http:' + photos[0].get_attribute('data-zoom-src')
            main_photo = self.get_photo(main_photo_url, str(article_num) + '_0.webp')

            other_photo = []
            for j in range(1, len(photos)):
                other_photo_url = 'http:' + photos[j].get_attribute('data-zoom-src')
                other_photo.append(self.get_photo(other_photo_url, str(article_num) + f'_{j}.webp'))
            other_photo = ','.join(other_photo)

            sizes = [j.text.split('\n')[0] for j in self.driver.find_elements(By.CLASS_NAME, 'size-container')]
            sizes = list(filter(lambda x: x != '', sizes))

            for size in sizes:
                c += 1
                article = 'COS_' + article_num + '_' + size
                text = ''
                for i in self.translate(description).split('\n'):
                    text += i + '","'
                rich = self.RICH.format(name, text, article_num)

                self.COLUMNS['№'] = c
                self.COLUMNS['Артикул*'] = article
                self.COLUMNS['Название товара'] = name
                self.COLUMNS['Цена, руб.*'] = price
                self.COLUMNS['Ссылка на главное фото*'] = main_photo
                self.COLUMNS['Ссылки на дополнительные фото'] = other_photo
                self.COLUMNS['Объединить на одной карточке*'] = article_num[:-3]
                self.COLUMNS['Цвет товара*'] = self.COLORS[color] if color in self.COLORS.keys() else 'разноцветный'
                self.COLUMNS['Название цвета'] = color
                self.COLUMNS['Материал'] = material
                self.COLUMNS['Состав материала'] = material
                self.COLUMNS['Таблица размеров JSON'] = self.TABLE_OF_SIZES
                self.COLUMNS['Rich-контент JSON'] = rich
                try:
                    self.COLUMNS['Российский размер*'] = self.SIZES[size.upper()]
                except:
                    self.COLUMNS['Российский размер*'] = 'Bad size'  # Если размера нету в таблице размеров
                self.COLUMNS['Размер производителя'] = size
                self.COLUMNS["Страна-изготовитель"] = self.translate(creator)

                self.result.append(self.COLUMNS.copy())

    def gPriceDict(self, key):
        return float(PRICE_TABLE[key])

    def get_price(self, eur_price):
        cost_price = (float(eur_price) * self.gPriceDict("КОЭФ_КОНВЕРТАЦИИ") * self.gPriceDict(
            'КУРС_EUR_RUB')) + (self.DELIVERY_PRICE * self.gPriceDict('КУРС_БЕЛ.РУБ_РУБ') * self.gPriceDict(
            'КУРС_EUR_БЕЛ.РУБ'))
        final_price = ((cost_price + self.gPriceDict('СРЕД_ЦЕН_ДОСТАВКИ')) * self.gPriceDict('НАЦЕНКА')) / (
                    1 - self.gPriceDict('ПРОЦЕНТЫ_ОЗОН') - self.gPriceDict('ПРОЦЕНТЫ_НАЛОГ') - self.gPriceDict('ПРОЦЕНТЫ_ЭКВАЙРИНГ'))

        if final_price > 20000:
            final_price = (final_price // 1000 + 1) * 1000 - 10
        elif final_price > 10000:
            if final_price % 1000 >= 500:
                final_price = (final_price // 1000) * 1000 + 990
            else:
                final_price = (final_price // 1000) * 1000 + 490
        else:
            final_price = (final_price // 100 + 1) * 100 - 10
        return final_price

    def check_exists_by_xpath(self, xpath):
        try:
            self.driver.find_element(By.XPATH, xpath)
        except NoSuchElementException:
            return False
        return True

    def get_photo(self, url, name):
        r = requests.get(url, stream=True)
        if r.status_code == 200:
            with open(SAVE_PHOTO_PATH + name, 'wb') as f:
                r.raw.decode_content = True
                shutil.copyfileobj(r.raw, f)
            return 'http://' + HOST + '/COS_parser/' + SAVE_PHOTO_PATH + name
        else:
            return 'Bad photo'

    def translate(self, text):
        translator = Translator()
        while True:
            try:
                result = translator.translate(text, dest='ru')
                return result.text
            except:
                pass

    def save(self, result):
        wb = load_workbook(filename=f'{self.data[CATEGORIE]["folder_path"]}/example.xlsx')
        ws = wb['Шаблон для поставщика']
        alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
        cols = []
        for col in alphabet:
            value = ws[col + '2'].value
            if value:
                cols.append(value)
        for col1 in alphabet:
            for col2 in alphabet:
                value = ws[col1 + col2 + '2'].value
                if value:
                    cols.append(value)

        for row in range(len(result)):
            for col in range(len(cols)):
                if cols[col] not in result[row]:
                    ws.cell(row=4 + row, column=1 + col).value = ''
                else:
                    ws.cell(row=4 + row, column=1 + col).value = result[row][cols[col]]

        wb.save(SAVE_XLSX_PATH + CATEGORIE + f"_{datetime.now()}.xlsx".replace(':', '.'))

    def sort_result(self):
        self.result.sort(key=lambda x: x['Артикул*'])
        for i in range(len(self.result)):
            self.result[i]['№'] = i + 1

    def load_settings(self):
        with open('settings.json', 'r', encoding='utf-8') as f:
            self.data = json.load(f)
        self.CATEGORIE_URL = self.data[CATEGORIE]['url']
        self.DELIVERY_PRICE = int(self.data[CATEGORIE]["ЦЕНА_ДОСТАВКИ_В_КАТЕГОРИИ"])
        self.COLUMNS = self.load_module('columns').COLUMNS
        self.RICH = self.load_module('rich').RICH
        self.SIZES = self.load_module('sizes').SIZES
        self.TABLE_OF_SIZES = self.load_module('table_of_sizes').TABLE_OF_SIZES
        self.MATERIALS = MATERIALS
        self.COLORS = COLORS

    def load_module(self, name):
        spec = importlib.util.spec_from_file_location(name, self.data[CATEGORIE]['folder_path'] + '/' + name + '.py')
        foo = importlib.util.module_from_spec(spec)
        sys.modules[name] = foo
        spec.loader.exec_module(foo)
        return foo

    def start(self):
        try:
            self.load_settings()
            print('--- START PARSING ---')
            self.parse()
            print('--- END PARSING ---')
        except Exception as e:
            error = self.driver.current_url + '\n' + traceback.format_exc() + '\n'
            print(error)
            with open('log.log', 'a') as f:
                f.write(error)
            with open('last.html', 'w') as f:
                f.write(self.driver.page_source)
        finally:
            self.sort_result()
            self.save(self.result)

            self.driver.close()
            self.driver.quit()


def main():
    parser = Parser()
    parser.start()


if __name__ == '__main__':
    if 'photo' not in os.listdir():
        os.mkdir('photo')
    if 'xlsx' not in os.listdir():
        os.mkdir('xlsx')
    if 'log.log' not in os.listdir():
        file = open('log.log', 'w')
        file.close()
    main()

